import pytest
import io
import openpyxl
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APIClient

from api.products.factories import ProductFactory
from api.comments.factories import CommentFactory


@pytest.fixture
def api_client():
    return APIClient()


@pytest.mark.django_db
class TestDynamicReportView:

    def test_generate_excel_products_success(self, api_client):
        """
        Arrange: Crear productos en la BD.
        Act: Solicitar reporte tipo 'products' en formato 'excel'.
        Assert: Verificar Content-Type y que el Excel tenga las columnas correctas.
        """
        ProductFactory.create_batch(5)
        url = reverse('reports:dynamic_report')
        params = {'type': 'products', 'formate': 'excel'}

        print("\nTest 1: Solicitar reporte tipo 'products' en formato 'excel'")

        response = api_client.get(url, params)

        print(f"DEBUG: {response.status_code}")

        assert response.status_code == status.HTTP_200_OK
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        excel_file = io.BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        assert sheet.max_row == 6
        assert sheet.cell(row=1, column=1).value == 'Nombre'

    def test_generate_pdf_comments_success(self, api_client):
        """
        Arrange: Crear comentarios en la BD.
        Act: Solicitar reporte tipo 'comments' en formato 'pdf'.
        Assert: Verificar que la respuesta sea un PDF válido.
        """
        CommentFactory.create_batch(3)
        url = reverse('reports:dynamic_report')
        params = {'type': 'comments', 'formate': 'pdf'}

        print("\nTest 2: Solicitar reporte tipo 'comments' en formato 'pdf'")

        response = api_client.get(url, params)

        print(f"DEBUG: Respuesta: {response.status_code}")

        assert response.status_code == status.HTTP_200_OK
        assert response['Content-Type'] == 'application/pdf'
        assert response.content.startswith(b'%PDF')
        assert f'reporte_comentarios.pdf' in response['Content-Disposition']

    def test_invalid_parameters_returns_400(self, api_client):
        """
        Arrange: URL sin parámetros.
        Act: Realizar petición GET.
        Assert: Retornar 400 Bad Request.
        """
        url = reverse('reports:dynamic_report')

        print("\nTest 3: Realizar petición GET sin parámetros")

        response = api_client.get(url)

        print(f"DEBUG: Respuesta: {response.status_code}")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
